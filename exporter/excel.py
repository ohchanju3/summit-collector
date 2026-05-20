import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import List, Dict


# 엑셀 컬럼 순서 및 한글 헤더
COLUMN_MAP = {
    "date": "날짜",
    "participants": "참가자",
    "location": "장소",
    "summary": "내용 요약",
    "title": "기사 제목",
    "source": "언론사",
    "source_country": "출처 국가",
    "url": "링크",
    "extraction_method": "추출 방식",
}


def save_to_excel(data: List[Dict], output_dir: str = "data") -> str:
    """처리된 데이터를 엑셀 파일로 저장"""
    Path(output_dir).mkdir(exist_ok=True)

    df = pd.DataFrame(data)[list(COLUMN_MAP.keys())]
    df = df.rename(columns=COLUMN_MAP)

    # 플레이스홀더 값 제거
    placeholders = {"City, Country", "City", "Country", "Name1, Name2", "Name1"}
    df["장소"] = df["장소"].apply(lambda x: "" if str(x).strip() in placeholders else x)
    df["참가자"] = df["참가자"].apply(lambda x: "" if str(x).strip() in placeholders else x)

    df = df.sort_values("날짜", ascending=False).reset_index(drop=True)

    filename = f"summits_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = Path(output_dir) / filename

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="정상회담")

        ws = writer.sheets["정상회담"]

        # 헤더 굵게
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 컬럼 너비 자동 조정
        col_widths = {"날짜": 12, "참가자": 30, "장소": 20, "내용 요약": 60,
                      "기사 제목": 60, "언론사": 20, "출처 국가": 12, "링크": 50, "추출 방식": 20}
        for i, col_name in enumerate(df.columns, 1):
            col_letter = ws.cell(1, i).column_letter
            ws.column_dimensions[col_letter].width = col_widths.get(col_name, 20)

    print(f"저장 완료: {filepath}  ({len(df)}행)")
    return str(filepath)
