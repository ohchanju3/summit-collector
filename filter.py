"""
기존 수집된 엑셀 파일에서 실제 정상회담 기사만 필터링하는 후처리 스크립트
Usage: python filter.py data/summits_20260519_121204.xlsx
"""
import sys
import json
import ollama
import pandas as pd
from pathlib import Path
from datetime import datetime
from config import LLM_MODEL


def is_summit_article(row: pd.Series) -> bool:
    """
    Llama 3로 해당 행이 실제 정상회담 기사인지 판단
    """
    def clean(val):
        s = str(val).strip()
        return "" if s in ("nan", "None", "") else s

    text = f"Title: {clean(row['기사 제목'])}\nParticipants: {clean(row['참가자'])}\nLocation: {clean(row['장소'])}\nSummary: {clean(row['내용 요약'])}"

    system_instruction = (
        "You are an expert in international diplomacy. "
        "Determine whether the given article is about an actual bilateral or multilateral summit meeting between heads of state or government. "
        "A valid summit: leaders (presidents, prime ministers, kings, etc.) of two or more countries actually met in person or virtually for official diplomatic purposes. "
        "NOT a valid summit: "
        "- one leader issuing a warning, statement, or demand to another country without meeting "
        "- news analysis or opinion pieces mentioning past summits "
        "- articles about a future/upcoming meeting that has not yet taken place "
        "- sports, business, or NGO events without heads of state "
        "- military conflicts, natural disasters, or unrelated political news "
        "Output strictly as JSON: {\"is_summit\": true} or {\"is_summit\": false}"
    )

    try:
        response = ollama.chat(
            model=LLM_MODEL,
            messages=[
                {'role': 'system', 'content': system_instruction},
                {'role': 'user', 'content': text}
            ],
            options={'temperature': 0.0, 'num_thread': 4},
            format='json'
        )
        result = json.loads(response['message']['content'])
        return bool(result.get("is_summit", False))
    except Exception as e:
        print(f"   [LLM Error] → {e}")
        return True  # 에러 시 보수적으로 포함


def main():
    if len(sys.argv) < 2:
        print("Usage: python filter.py <excel_file>")
        print("Example: python filter.py data/summits_20260519_121204.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    df = pd.read_excel(input_path)
    total = len(df)
    print(f"총 {total}개 행 로드 완료 → 정상회담 여부 판단 시작\n")

    results = []
    kept = 0

    for idx, row in df.iterrows():
        title = str(row['기사 제목'])[:40]
        verdict = is_summit_article(row)
        label = "O" if verdict else "X"
        print(f"  [{idx+1}/{total}] [{label}] {title}...")

        if verdict:
            kept += 1
            results.append(row)

    filtered_df = pd.DataFrame(results).reset_index(drop=True)

    # 저장
    output_dir = Path(input_path).parent
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = output_dir / f"summits_filtered_{timestamp}.xlsx"

    from openpyxl.styles import Font, PatternFill, Alignment
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        filtered_df.to_excel(writer, index=False, sheet_name="정상회담")
        ws = writer.sheets["정상회담"]

        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        col_widths = {"날짜": 12, "참가자": 30, "장소": 20, "내용 요약": 60,
                      "기사 제목": 60, "언론사": 20, "출처 국가": 12, "링크": 50, "추출 방식": 20}
        for i, col_name in enumerate(filtered_df.columns, 1):
            col_letter = ws.cell(1, i).column_letter
            ws.column_dimensions[col_letter].width = col_widths.get(col_name, 20)

    print(f"\n완료: {total}개 → {kept}개 ({total - kept}개 제거)")
    print(f"저장: {output_path}")


if __name__ == "__main__":
    main()
